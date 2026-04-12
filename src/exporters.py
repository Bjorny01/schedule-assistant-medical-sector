"""
Output exporters for the Medical Staff Rostering System.

Produces:
  • One .ics calendar file per staff member
  • One .xlsx admin overview with colour-coded shift grid
"""
from __future__ import annotations

from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path
from zoneinfo import ZoneInfo

from .models import ParsedInputs, Schedule, ShiftAssignment, ShiftType, Staff

TIMEZONE = ZoneInfo("Europe/Stockholm")

# Shift start/end times (local)
SHIFT_TIMES: dict[ShiftType, tuple[time, time]] = {
    ShiftType.DAY:     (time(7, 0),  time(15, 0)),
    ShiftType.EVENING: (time(15, 0), time(23, 0)),
    ShiftType.NIGHT:   (time(23, 0), time(7, 0)),   # ends next calendar day
}

# Colour fills for Excel (openpyxl PatternFill hex)
SHIFT_COLOURS: dict[ShiftType, str] = {
    ShiftType.DAY:     "BDD7EE",   # light blue
    ShiftType.EVENING: "FFE699",   # light amber
    ShiftType.NIGHT:   "7030A0",   # purple (white text)
}
SHIFT_FONT_DARK = {"DAY", "EVENING"}  # shifts with dark text
OFF_COLOUR = "F2F2F2"


# ---------------------------------------------------------------------------
# ICS export
# ---------------------------------------------------------------------------

def export_ics_files(schedule: Schedule, inputs: ParsedInputs, output_dir: Path) -> None:
    """Write one .ics file per staff member into output_dir."""
    try:
        from icalendar import Calendar, Event
    except ImportError:
        print("[exporters] icalendar not installed — skipping .ics export. "
              "Run: pip install icalendar")
        return

    output_dir.mkdir(parents=True, exist_ok=True)

    for staff in inputs.staff:
        cal = Calendar()
        cal.add("prodid", "-//Med-3 Rostering System//EN")
        cal.add("version", "2.0")
        cal.add("calscale", "GREGORIAN")
        cal.add("x-wr-calname", f"Med-3 — {staff.name}")
        cal.add("x-wr-timezone", "Europe/Stockholm")

        for assignment in schedule.get_staff_assignments(staff.name):
            event = _make_ical_event(assignment)
            cal.add_component(event)

        safe_name = staff.name.lower().replace(" ", "_")
        path = output_dir / f"{safe_name}.ics"
        path.write_bytes(cal.to_ical())
        print(f"[exporters] Written {path}")


def _make_ical_event(assignment: ShiftAssignment):
    from icalendar import Event, vDatetime, vText

    event = Event()
    shift_type = assignment.shift_type
    start_time, end_time = SHIFT_TIMES[shift_type]
    shift_label = shift_type.value.capitalize()

    dt_start = datetime.combine(assignment.date, start_time, tzinfo=TIMEZONE)
    if shift_type == ShiftType.NIGHT:
        dt_end = datetime.combine(assignment.date + timedelta(days=1), end_time, tzinfo=TIMEZONE)
    else:
        dt_end = datetime.combine(assignment.date, end_time, tzinfo=TIMEZONE)

    event.add("summary",     f"{shift_label} shift — Med-3")
    event.add("dtstart",     dt_start)
    event.add("dtend",       dt_end)
    event.add("description", (
        f"Staff: {assignment.staff_name}\n"
        f"Shift: {shift_label} "
        f"({start_time.strftime('%H:%M')}–"
        f"{end_time.strftime('%H:%M')}"
        f"{' +1d' if shift_type == ShiftType.NIGHT else ''})\n"
        f"Ward: Medicin Avdelning 3 (Med-3)"
    ))
    event.add("location", "Medicin Avdelning 3, Länssjukhuset Exempel")
    event.add("status",    "CONFIRMED")

    return event


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def export_excel(schedule: Schedule, inputs: ParsedInputs, output_dir: Path) -> None:
    """Write an admin overview .xlsx file into output_dir."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[exporters] openpyxl not installed — skipping .xlsx export. "
              "Run: pip install openpyxl")
        return

    output_dir.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Schedule Overview"

    num_days = schedule.num_weeks * 7
    dates = [schedule.start_date + timedelta(days=d) for d in range(num_days)]

    # ------------------------------------------------------------------
    # Headers
    # ------------------------------------------------------------------

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=9)
    week_fill = PatternFill(fill_type="solid", fgColor="2E75B6")
    week_font = Font(bold=True, color="FFFFFF", size=8)

    # Row 1: week labels
    ws.cell(row=1, column=1, value="Staff \\ Date")
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).font = header_font
    ws.cell(row=1, column=1).border = border
    ws.column_dimensions["A"].width = 22

    for d, day in enumerate(dates):
        col = d + 2
        if d % 7 == 0:
            week_num = d // 7 + 1
            ws.cell(row=1, column=col, value=f"W{week_num}")
            ws.cell(row=1, column=col).fill = week_fill
            ws.cell(row=1, column=col).font = week_font
            ws.cell(row=1, column=col).border = border

    # Row 2: date headers (Mon dd/mm format)
    for d, day in enumerate(dates):
        col = d + 2
        label = day.strftime("%a\n%d/%m")
        cell = ws.cell(row=2, column=col, value=label)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        is_weekend = day.weekday() >= 5
        cell.fill = PatternFill(fill_type="solid", fgColor="D6E4F0" if not is_weekend else "FDEBD0")
        cell.font = Font(bold=True, size=8)
        ws.column_dimensions[get_column_letter(col)].width = 6

    ws.row_dimensions[2].height = 28

    # ------------------------------------------------------------------
    # Data rows — one per staff member
    # ------------------------------------------------------------------

    shift_fills = {
        st: PatternFill(fill_type="solid", fgColor=col)
        for st, col in SHIFT_COLOURS.items()
    }
    off_fill = PatternFill(fill_type="solid", fgColor=OFF_COLOUR)

    for row_idx, staff in enumerate(inputs.staff):
        row = row_idx + 3
        assignments_by_date = {
            a.date: a for a in schedule.get_staff_assignments(staff.name)
        }

        # Staff name cell
        name_cell = ws.cell(
            row=row,
            column=1,
            value=f"{staff.name}\n{staff.role.value.capitalize()} · {int(staff.contract_pct*100)}%",
        )
        name_cell.font = Font(bold=True, size=9)
        name_cell.fill = PatternFill(fill_type="solid", fgColor="DEEAF1")
        name_cell.border = border
        name_cell.alignment = Alignment(vertical="center", wrap_text=True)

        for d, day in enumerate(dates):
            col = d + 2
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(size=9)

            assignment = assignments_by_date.get(day)
            if assignment:
                cell.value = SHIFT_SYMBOLS_EXCEL[assignment.shift_type]
                cell.fill = shift_fills[assignment.shift_type]
                if assignment.shift_type == ShiftType.NIGHT:
                    cell.font = Font(size=9, color="FFFFFF", bold=True)
            else:
                cell.value = ""
                cell.fill = off_fill

        ws.row_dimensions[row].height = 28

    # ------------------------------------------------------------------
    # Summary row — totals per staff vs contract
    # ------------------------------------------------------------------

    summary_row = len(inputs.staff) + 3
    ws.cell(row=summary_row, column=1, value="TOTAL SHIFTS").font = Font(bold=True, size=9)

    for row_idx, staff in enumerate(inputs.staff):
        total = schedule.total_shifts_for(staff.name)
        target = round(staff.target_shifts_per_week * schedule.num_weeks)
        row = row_idx + 3
        summary_cell = ws.cell(row=row, column=num_days + 2,
                               value=f"{total}/{target}")
        summary_cell.font = Font(
            bold=True, size=9,
            color="C00000" if total < target - 2 else "375623",
        )
        summary_cell.border = border
        summary_cell.alignment = Alignment(horizontal="center")

    ws.cell(row=2, column=num_days + 2, value="Sched/Target").font = Font(bold=True, size=8)
    ws.column_dimensions[get_column_letter(num_days + 2)].width = 12

    # ------------------------------------------------------------------
    # Legend sheet
    # ------------------------------------------------------------------

    legend_ws = wb.create_sheet("Legend")
    legend_ws.column_dimensions["A"].width = 18
    legend_ws.column_dimensions["B"].width = 30

    legend_data = [
        ("Symbol", "Meaning", None),
        ("D", "Day shift  07:00–15:00", SHIFT_COLOURS[ShiftType.DAY]),
        ("E", "Evening shift  15:00–23:00", SHIFT_COLOURS[ShiftType.EVENING]),
        ("N", "Night shift  23:00–07:00", SHIFT_COLOURS[ShiftType.NIGHT]),
        ("(empty)", "Day off", OFF_COLOUR),
    ]

    for i, (sym, desc, colour) in enumerate(legend_data, start=1):
        a = legend_ws.cell(row=i, column=1, value=sym)
        b = legend_ws.cell(row=i, column=2, value=desc)
        if colour:
            fill = PatternFill(fill_type="solid", fgColor=colour)
            a.fill = fill
            b.fill = fill
        if i == 1:
            a.font = Font(bold=True)
            b.font = Font(bold=True)

    path = output_dir / "schedule_overview.xlsx"
    wb.save(path)
    print(f"[exporters] Written {path}")


SHIFT_SYMBOLS_EXCEL = {
    ShiftType.DAY:     "D",
    ShiftType.EVENING: "E",
    ShiftType.NIGHT:   "N",
}
